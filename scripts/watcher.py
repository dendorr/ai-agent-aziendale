import sys
import os
sys.path.append(os.path.expanduser("~/ai-agent"))

from config.config import WATCH_FOLDER, CHROMA_DB_PATH, SUPPORTED_EXTENSIONS
from pathlib import Path
import chromadb
import pdfplumber
import openpyxl
import ezdxf
import time
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

client = chromadb.PersistentClient(path=CHROMA_DB_PATH)
collection = client.get_or_create_collection("documenti_aziendali")

def leggi_pdf(filepath):
    import fitz
    testo = ""
    doc = fitz.open(filepath)
    for pagina in doc:
        testo += pagina.get_text() + "\n"
    doc.close()
    return testo

def chunk_testo(testo, chunk_size=500, overlap=50):
    parole = testo.split()
    chunks = []
    i = 0
    while i < len(parole):
        chunk = " ".join(parole[i:i+chunk_size])
        chunks.append(chunk)
        i += chunk_size - overlap
    return chunks

def leggi_excel(filepath):
    testo = ""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    for foglio in wb.sheetnames:
        ws = wb[foglio]
        testo += f"Foglio: {foglio}\n"
        for riga in ws.iter_rows(values_only=True):
            riga_pulita = [str(c) for c in riga if c is not None]
            if riga_pulita:
                testo += " | ".join(riga_pulita) + "\n"
    return testo

def leggi_dxf(filepath):
    testo = ""
    doc = ezdxf.readfile(filepath)
    msp = doc.modelspace()
    for entita in msp:
        if entita.dxftype() == "TEXT":
            testo += entita.dxf.text + "\n"
        elif entita.dxftype() == "MTEXT":
            testo += entita.plain_mtext() + "\n"
    return testo

def leggi_file(filepath):
    ext = Path(filepath).suffix.lower()
    try:
        if ext == ".pdf":
            return leggi_pdf(filepath)
        elif ext in [".xlsx", ".xls"]:
            return leggi_excel(filepath)
        elif ext == ".dxf":
            return leggi_dxf(filepath)
        elif ext in [".txt", ".md", ".csv", ".svg"]:
            with open(filepath, "r", errors="ignore") as f:
                return f.read()
        else:
            return f"File {ext} rilevato: {Path(filepath).name}"
    except Exception as e:
        return f"Errore nella lettura: {e}"

def indicizza_file(filepath):
    ext = Path(filepath).suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        return
    print(f"[+] Nuovo file rilevato: {Path(filepath).name}")
    testo = leggi_file(filepath)
    if testo:
        chunks = chunk_testo(testo)
        for i, chunk in enumerate(chunks):
            collection.upsert(
                documents=[chunk],
                ids=[f"{str(filepath)}__chunk{i}"],
                metadatas=[{"filename": Path(filepath).name, "path": str(filepath), "chunk": i}]
            )
        print(f"[✓] Indicizzato: {Path(filepath).name} → {len(chunks)} chunks")

def rimuovi_file(filepath):
    try:
        collection.delete(ids=[str(filepath)])
        print(f"[-] Rimosso dal database: {Path(filepath).name}")
    except:
        pass

class FileHandler(FileSystemEventHandler):
    def on_created(self, event):
        if not event.is_directory:
            time.sleep(1)
            indicizza_file(event.src_path)

    def on_modified(self, event):
        if not event.is_directory:
            time.sleep(1)
            indicizza_file(event.src_path)

    def on_deleted(self, event):
        if not event.is_directory:
            rimuovi_file(event.src_path)

    def on_moved(self, event):
        if not event.is_directory:
            rimuovi_file(event.src_path)
            indicizza_file(event.dest_path)

if __name__ == "__main__":
    print(f"[*] Watcher avviato su: {WATCH_FOLDER}")
    print(f"[*] Documenti già nel database: {collection.count()}")
    print(f"[*] Controllo ogni 10 secondi... (Ctrl+C per fermare)")

    file_visti = set()

    while True:
        try:
            cartella = Path(WATCH_FOLDER)
            for filepath in cartella.rglob("*"):
                if filepath.is_file() and filepath.suffix.lower() in SUPPORTED_EXTENSIONS:
                    if str(filepath) not in file_visti:
                        file_visti.add(str(filepath))
                        indicizza_file(str(filepath))
            time.sleep(10)
        except KeyboardInterrupt:
            print("\n[*] Watcher fermato.")
            break
        except Exception as e:
            print(f"[!] Errore: {e}")
            time.sleep(10)
