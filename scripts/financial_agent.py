"""
FINANCIAL AGENT v10 (async) — Universal financial document intelligence

Zero hardcoded assumptions su struttura, contenuto o dominio dei file.
Multi-model pipeline:
  - ROUTING_MODEL (LLM_MODEL_FAST) → seleziona tabelle, genera SQL
  - ANSWER_MODEL  (LLM_MODEL_MAIN) → produce la risposta finale in italiano

Tutte le funzioni di lettura Excel/CSV/PDF sono sincrone (CPU-bound).
Le chiamate LLM sono async via llm_client.
SQLite locale resta sincrono (operazioni rapide su DB locale).

Excel:
  - Auto-detect riga header (gestisce title rows prima delle intestazioni)
  - Date convertite in stringhe ISO
  - Merged cells: master value propagato alle figlie
  - Color sampling sulle prime 500 righe

PDF: markitdown → fitz → pdfplumber (fallback a tre livelli).
"""

import sys
import os
import json
import sqlite3
import re
import hashlib
import asyncio
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config.config import (
    CHROMA_PATHS, LLM_MODEL_MAIN, LLM_MODEL_FAST,
    CHUNK_SIZE, CHUNK_OVERLAP, EXTENSIONS, MEMORY_PATH,
)

import chromadb
import fitz
import semantic_analyzer as analyzer
from llm_client import chat_complete, chat_complete_stream, chat_complete_json

client     = chromadb.PersistentClient(path=CHROMA_PATHS["financial"])
collection = client.get_or_create_collection("financial")

MEMORY_FILE = Path(MEMORY_PATH) / "financial_memory.json"
GENERIC_DB  = Path(MEMORY_PATH) / "financial_files.db"

# ── Models ────────────────────────────────────────────────────────────────────
ROUTING_MODEL = LLM_MODEL_FAST
ANSWER_MODEL  = LLM_MODEL_MAIN

# RGB trattati come "no background" (default/bianco/nero)
_IGNORE_RGB = {"00000000", "FFFFFFFF", "00FFFFFF", "FF000000"}


# ── Helper low-level ──────────────────────────────────────────────────────────

def sanitize_col(name: str) -> str:
    """Converte una stringa qualsiasi in un nome colonna SQLite valido (max 50)."""
    s = re.sub(r"[^a-zA-Z0-9_]", "_", str(name).strip()).strip("_")
    return (s or "col")[:50]


def table_name_for(filepath: str, sheet: str = "") -> str:
    """Genera un nome tabella SQLite deterministico e collision-free."""
    h = hashlib.md5(f"{filepath}_{sheet}".encode()).hexdigest()[:8]
    stem = re.sub(r"[^a-zA-Z0-9]", "_", Path(filepath).stem[:20]).strip("_")
    return f"f_{stem}_{h}".lower()


def query_db(db_path, sql: str, params=()):
    """Esegue una query read-only su SQLite. Ritorna lista di dict."""
    if not Path(db_path).exists():
        return []
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        return [dict(r) for r in conn.execute(sql, params).fetchall()]
    except Exception as e:
        return [{"error": str(e)}]
    finally:
        conn.close()


# ── Generic SQLite (un DB, una tabella per file+sheet) ────────────────────────

def _open_generic_db():
    GENERIC_DB.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(GENERIC_DB)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS _file_schemas (
            schema_id  TEXT PRIMARY KEY,
            filepath   TEXT,
            filename   TEXT,
            sheet      TEXT,
            table_name TEXT,
            col_names  TEXT,
            col_raw    TEXT,
            col_types  TEXT,
            col_stats  TEXT,
            color_info TEXT,
            row_count  INTEGER,
            indexed_at TEXT
        )
    """)
    conn.commit()
    return conn


# ── Excel cell helpers ────────────────────────────────────────────────────────

def _cell_value(cell):
    """
    Estrae un valore Python pulito da una cella openpyxl.
    - Formule  → None  (data_only=True le risolve)
    - datetime → stringa ISO YYYY-MM-DD
    - Altro    → valore raw
    """
    import datetime as _dt
    v = cell.value
    if v is None:
        return None
    if isinstance(v, str) and v.strip().startswith("="):
        return None
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.strftime("%Y-%m-%d")
    return v


def _cell_rgb(cell):
    """Hex RGB di sfondo, None se default/bianco/nero."""
    try:
        rgb = cell.fill.fgColor.rgb
        if rgb and rgb.upper() not in _IGNORE_RGB:
            return rgb.upper()
    except Exception:
        pass
    return None


# ── Excel structure helpers ───────────────────────────────────────────────────

def _detect_header_row(ws, max_scan: int = 5) -> int:
    """Scansiona le prime righe e ritorna l'indice (1-based) della riga
    con più celle non-null (= probabile riga header)."""
    best_row, best_count = 1, 0
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan), 1):
        count = sum(1 for c in row if c.value is not None)
        if count > best_count:
            best_count, best_row = count, i
    return best_row


def _build_merge_map(ws) -> dict:
    """{(row, col): value} per ogni cella non-master in ogni merge range."""
    merge_map = {}
    try:
        for rng in ws.merged_cells.ranges:
            master_val = ws.cell(row=rng.min_row, column=rng.min_col).value
            for row in range(rng.min_row, rng.max_row + 1):
                for col in range(rng.min_col, rng.max_col + 1):
                    if (row, col) != (rng.min_row, rng.min_col):
                        merge_map[(row, col)] = master_val
    except Exception:
        pass
    return merge_map


# ── Type detection colonne ────────────────────────────────────────────────────

def _detect_col_stats(df) -> tuple:
    """Auto-detect type (numeric/text/empty) e calcola statistiche."""
    import pandas as pd
    col_types, col_stats = {}, {}

    for col in df.columns:
        series = df[col].dropna()
        if series.empty:
            col_types[col] = "empty"
            continue

        num = pd.to_numeric(series, errors="coerce").dropna()
        if len(num) / len(series) >= 0.7:
            col_types[col] = "numeric"
            col_stats[col] = {
                "sum":     round(float(num.sum()),  4),
                "mean":    round(float(num.mean()), 4),
                "min":     round(float(num.min()),  4),
                "max":     round(float(num.max()),  4),
                "count":   int(num.count()),
                "nonzero": int((num != 0).sum()),
            }
        else:
            col_types[col] = "text"
            vc = series.astype(str).value_counts()
            col_stats[col] = {
                "unique":     int(series.nunique()),
                "top_values": vc.head(10).index.tolist(),
                "count":      int(series.count()),
            }

    return col_types, col_stats


# ── Sheet processor ───────────────────────────────────────────────────────────

def _process_sheet(ws, sname: str, color_sample: int = 500):
    """Processa un singolo sheet openpyxl. Ritorna data dict o None."""
    import pandas as pd

    if ws.max_row is None or ws.max_row < 2:
        return None

    hdr_idx = _detect_header_row(ws)
    hdr_row = list(ws.iter_rows(min_row=hdr_idx, max_row=hdr_idx))[0]
    raw_hdrs = [c.value for c in hdr_row]

    if not any(h for h in raw_hdrs if h is not None):
        return None

    n_cols = len(raw_hdrs)

    seen, san_hdrs = {}, []
    for i, h in enumerate(raw_hdrs):
        base = sanitize_col(h) if h is not None else f"col_{i}"
        cnt = seen.get(base, 0)
        seen[base] = cnt + 1
        san_hdrs.append(f"{base}_{cnt}" if cnt else base)

    merge_map = _build_merge_map(ws)
    rows = []
    color_info = {}
    color_rows = 0

    for ri, row in enumerate(ws.iter_rows(min_row=hdr_idx + 1), hdr_idx + 1):
        vals = []
        for ci, cell in enumerate(row[:n_cols]):
            val = merge_map.get((ri, cell.column), _cell_value(cell))
            vals.append(val)

            if color_rows < color_sample and ci < len(san_hdrs):
                rgb = _cell_rgb(cell)
                if rgb:
                    col = san_hdrs[ci]
                    if col not in color_info:
                        color_info[col] = {}
                    color_info[col][rgb] = color_info[col].get(rgb, 0) + 1

        rows.append(vals + [None] * max(0, n_cols - len(vals)))
        if ri - hdr_idx <= color_sample:
            color_rows += 1

    if not rows:
        return None

    import pandas as pd
    df = pd.DataFrame(rows, columns=san_hdrs)
    col_types, col_stats = _detect_col_stats(df)

    return {
        "df":         df,
        "col_names":  san_hdrs,
        "col_raw":    [str(h) if h is not None else "" for h in raw_hdrs],
        "col_types":  col_types,
        "col_stats":  col_stats,
        "color_info": color_info,
    }


def _sheet_summary_lines(sname: str, data: dict) -> list:
    """Righe leggibili di summary per un sheet (salvate nei chunk ChromaDB)."""
    san_hdrs = data["col_names"]
    raw_hdrs = data["col_raw"]
    col_types = data["col_types"]
    col_stats = data["col_stats"]
    color_info = data["color_info"]
    df = data["df"]

    raw_map = dict(zip(san_hdrs, raw_hdrs))
    raw_str = ", ".join(h for h in raw_hdrs if h)
    num_cols = [c for c in san_hdrs if col_types.get(c) == "numeric"]
    txt_cols = [c for c in san_hdrs if col_types.get(c) == "text"]

    lines = [f"[Foglio: {sname}] {len(df):,} righe | Colonne: {raw_str[:150]}"]

    if num_cols:
        lines.append("  Colonne numeriche:")
        for col in num_cols[:10]:
            s = col_stats[col]
            lines.append(
                f"    {raw_map.get(col, col)}: "
                f"sum={s['sum']:,.2f} | min={s['min']:,.2f} | "
                f"max={s['max']:,.2f} | n={s['count']}"
            )

    if txt_cols:
        lines.append("  Colonne testo:")
        for col in txt_cols[:10]:
            s = col_stats[col]
            top = ", ".join(str(v) for v in s["top_values"][:5])
            lines.append(
                f"    {raw_map.get(col, col)}: {s['unique']} valori unici (es: {top})"
            )

    if color_info:
        lines.append("  Pattern colori (sfondo celle — significato da inferire dal contesto):")
        for col, colors in list(color_info.items())[:6]:
            lbl = raw_map.get(col, col)
            for rgb, count in sorted(colors.items(), key=lambda x: -x[1])[:4]:
                hex_c = f"#{rgb[2:] if len(rgb) == 8 else rgb}"
                lines.append(f"    {lbl}: {count} celle con sfondo {hex_c}")

    return lines


# ── Excel / CSV readers ───────────────────────────────────────────────────────

def read_excel_smart(filepath, max_rows: int = 50_000, color_sample: int = 500):
    """Legge qualsiasi Excel. Single workbook open, auto-header, date, merge, colori."""
    import openpyxl
    p = Path(filepath)
    wb = openpyxl.load_workbook(filepath, data_only=True)

    sheets_data = {}
    summary_lines = [
        f"=== EXCEL: {p.name} ===",
        f"Fogli ({len(wb.sheetnames)}): {', '.join(wb.sheetnames[:30])}",
        "",
    ]

    for sname in wb.sheetnames:
        ws = wb[sname]
        try:
            data = _process_sheet(ws, sname, color_sample=color_sample)
        except Exception as e:
            summary_lines.append(f"[Foglio: {sname}] ERRORE: {e}")
            continue

        if data is None:
            continue

        sheets_data[sname] = data
        summary_lines.extend(_sheet_summary_lines(sname, data))
        summary_lines.append("")

    wb.close()
    return sheets_data, "\n".join(summary_lines)


def read_csv_smart(filepath, max_rows: int = 50_000):
    """Legge qualsiasi CSV con encoding auto-detect e schema dinamico."""
    import pandas as pd
    p = Path(filepath)
    df = None

    for enc in ["utf-8", "latin-1", "cp1252"]:
        try:
            df = pd.read_csv(filepath, encoding=enc, nrows=max_rows)
            break
        except UnicodeDecodeError:
            continue
        except Exception as e:
            return {}, f"[CSV error: {e}]"

    if df is None:
        return {}, f"[Impossibile leggere {p.name}]"

    seen, san_map = {}, {}
    for h in df.columns:
        base = sanitize_col(h)
        cnt = seen.get(base, 0)
        seen[base] = cnt + 1
        san_map[h] = f"{base}_{cnt}" if cnt else base

    raw_names = list(df.columns)
    df = df.rename(columns=san_map)
    col_types, col_stats = _detect_col_stats(df)

    sheets_data = {"CSV": {
        "df":         df,
        "col_names":  list(df.columns),
        "col_raw":    raw_names,
        "col_types":  col_types,
        "col_stats":  col_stats,
        "color_info": {},
    }}

    raw_map = dict(zip(list(df.columns), raw_names))
    num_cols = [c for c in df.columns if col_types.get(c) == "numeric"]
    txt_cols = [c for c in df.columns if col_types.get(c) == "text"]

    lines = [
        f"=== CSV: {p.name} ===",
        f"Righe: {len(df):,} | Colonne: {', '.join(raw_names[:30])}",
        "",
    ]
    if num_cols:
        lines.append("Colonne numeriche:")
        for col in num_cols[:10]:
            s = col_stats[col]
            lines.append(
                f"  {raw_map.get(col, col)}: "
                f"sum={s['sum']:,.2f} | min={s['min']:.2f} | max={s['max']:.2f}"
            )
    if txt_cols:
        lines.append("Colonne testo:")
        for col in txt_cols[:10]:
            s = col_stats[col]
            top = ", ".join(str(v) for v in s["top_values"][:5])
            lines.append(
                f"  {raw_map.get(col, col)}: {s['unique']} valori unici (es: {top})"
            )

    return sheets_data, "\n".join(lines)


# ── Generic DB builder ────────────────────────────────────────────────────────

def build_file_db(filepath, sheets_data) -> list:
    """Persiste ogni sheet nel DB SQLite generico. Una tabella per sheet."""
    import pandas as pd
    conn = _open_generic_db()
    p = Path(filepath)
    fhash = hashlib.md5(str(filepath).encode()).hexdigest()[:12]
    created = []

    for sname, data in sheets_data.items():
        df = data["df"]
        col_names = data["col_names"]
        col_raw = data["col_raw"]
        col_types = data["col_types"]
        col_stats = data["col_stats"]
        color_info = data.get("color_info", {})

        if df is None or df.empty:
            continue

        tname = table_name_for(filepath, sname)
        schema_id = f"{fhash}_{sname}"

        conn.execute(f'DROP TABLE IF EXISTS "{tname}"')
        col_defs = [
            f'"{col}" {"REAL" if col_types.get(col) == "numeric" else "TEXT"}'
            for col in col_names
        ]
        conn.execute(f'CREATE TABLE "{tname}" ({", ".join(col_defs)})')

        for col in col_names:
            if col_types.get(col) == "numeric":
                df[col] = pd.to_numeric(df[col], errors="coerce")
            else:
                df[col] = (df[col].fillna("")
                                  .astype(str)
                                  .str.strip()
                                  .replace({"nan": "", "None": "", "NaT": ""}))

        placeholders = ", ".join(["?"] * len(col_names))
        insert_sql = f'INSERT INTO "{tname}" VALUES ({placeholders})'
        batch = []
        for row in df.itertuples(index=False):
            vals = []
            for col, v in zip(col_names, row):
                if v is None or (isinstance(v, float) and pd.isna(v)):
                    vals.append(None)
                elif col_types.get(col) == "numeric":
                    try:
                        vals.append(float(v))
                    except Exception:
                        vals.append(None)
                else:
                    vals.append(str(v) if v is not None else "")
            batch.append(vals)
            if len(batch) >= 1000:
                conn.executemany(insert_sql, batch)
                batch = []
        if batch:
            conn.executemany(insert_sql, batch)

        for col in col_names:
            if col_types.get(col) == "text":
                if col_stats.get(col, {}).get("unique", 9999) < 500:
                    idx = f"idx_{tname[:18]}_{col[:12]}"
                    try:
                        conn.execute(
                            f'CREATE INDEX IF NOT EXISTS "{idx}" ON "{tname}"("{col}")'
                        )
                    except Exception:
                        pass

        conn.execute("""
            INSERT OR REPLACE INTO _file_schemas
            (schema_id, filepath, filename, sheet, table_name,
             col_names, col_raw, col_types, col_stats, color_info,
             row_count, indexed_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,datetime('now'))
        """, (
            schema_id, str(filepath), p.name, sname, tname,
            json.dumps(col_names,  ensure_ascii=False),
            json.dumps(col_raw,    ensure_ascii=False),
            json.dumps(col_types,  ensure_ascii=False),
            json.dumps(col_stats,  ensure_ascii=False),
            json.dumps(color_info, ensure_ascii=False),
            len(df),
        ))
        created.append((tname, sname, len(df)))

    conn.commit()
    conn.close()
    return created


# ── Schema retrieval ──────────────────────────────────────────────────────────

def get_all_file_schemas() -> list:
    """Tutti gli schemi indicizzati dal DB generico, con campi JSON parsati."""
    if not GENERIC_DB.exists():
        return []
    rows = query_db(GENERIC_DB, "SELECT * FROM _file_schemas ORDER BY indexed_at DESC")
    result = []
    for r in rows:
        s = dict(r)
        for key in ("col_names", "col_raw", "col_types", "col_stats", "color_info"):
            try:
                s[f"_{key}"] = json.loads(s.get(key) or "null") or {}
            except Exception:
                s[f"_{key}"] = {}
        result.append(s)
    return result


# ── Routing model (async) ─────────────────────────────────────────────────────

async def route_query(query: str, schemas: list) -> list:
    """Usa il routing model per selezionare tabelle e generare SQL."""
    if not schemas:
        return []

    schema_lines = []
    for s in schemas[:8]:
        col_names = s.get("_col_names") or []
        col_raw = s.get("_col_raw") or col_names
        col_types = s.get("_col_types") or {}
        raw_map = dict(zip(col_names, col_raw)) if len(col_names) == len(col_raw) else {}

        num_cols = [raw_map.get(c, c) for c in col_names if col_types.get(c) == "numeric"][:6]
        txt_cols = [raw_map.get(c, c) for c in col_names if col_types.get(c) == "text"][:6]

        schema_lines.append(
            f"  tabella={s['table_name']} | file={s['filename']} | "
            f"foglio={s['sheet']} | righe={s['row_count']} | "
            f"colonne_numeriche=[{', '.join(str(c) for c in num_cols)}] | "
            f"colonne_testo=[{', '.join(str(c) for c in txt_cols)}]"
        )

    system = "Sei un analista SQL esperto. Rispondi SOLO con JSON valido, niente testo extra."

    user = f"""TABELLE DISPONIBILI:
{chr(10).join(schema_lines)}

DOMANDA UTENTE: {query}

Genera SOLO un JSON array (massimo 3 query). Ogni elemento deve avere:
{{"table": "nome_tabella_esatto", "sql": "SELECT ... FROM \\"nome_tabella\\" ..."}}

Regole SQL obbligatorie:
- Nomi tabella e colonna ESATTAMENTE come indicati, tra doppi apici
- Per totali/somme: SUM("colonna")
- Per distribuzioni: GROUP BY "colonna" ORDER BY totale DESC
- Per ricerche testuali: WHERE "colonna" LIKE '%valore%'
- LIMIT 30 in ogni query
- Solo JSON valido, zero testo fuori dal JSON"""

    try:
        raw = await chat_complete_json(
            model=ROUTING_MODEL, system=system, user=user, temperature=0.0
        )
        match = re.search(r'\[.*?\]', raw, re.DOTALL)
        if match:
            parsed = json.loads(match.group())
            if isinstance(parsed, list):
                return parsed
    except Exception as e:
        print(f"  [routing] {e}", flush=True)

    return []


# ── Adaptive SQL context (async per via di route_query) ───────────────────────

async def adaptive_sql_context(query: str, schemas: list) -> str:
    """
    Costruisce contesto analitico ricco partendo dagli schemi SQLite.

    Phase 1 — Routing model genera SQL → eseguite localmente
    Phase 2 — Stats pre-calcolate + group-by per tutte le tabelle
    """
    if not schemas:
        return ""

    parts = []
    routed_tables = set()

    # ── Phase 1: routed SQL ──────────────────────────────────────────────────
    routed = await route_query(query, schemas)

    for item in routed:
        tname = item.get("table", "")
        sql = item.get("sql", "")
        if not tname or not sql:
            continue

        rows = query_db(GENERIC_DB, sql)
        if not rows or "error" in rows[0]:
            continue

        sch = next((s for s in schemas if s["table_name"] == tname), None)
        fname = sch["filename"] if sch else tname
        sheet = sch["sheet"] if sch else ""

        header = f"\n[Query → {fname}"
        if sheet and sheet != "CSV":
            header += f" | {sheet}"
        header += "]"
        parts.append(header)
        parts.append(f"SQL: {sql[:250]}")

        col_keys = list(rows[0].keys())
        parts.append(" | ".join(col_keys[:10]))
        for row in rows[:30]:
            parts.append("  " + " | ".join(str(row.get(k, "")) for k in col_keys[:10]))

        routed_tables.add(tname)

    # ── Phase 2: stats generiche ─────────────────────────────────────────────
    for sch in schemas[:6]:
        tname = sch["table_name"]
        fname = sch["filename"]
        sheet = sch["sheet"]
        col_names = sch.get("_col_names") or []
        col_raw = sch.get("_col_raw") or col_names
        col_types = sch.get("_col_types") or {}
        col_stats = sch.get("_col_stats") or {}
        color_info = sch.get("_color_info") or {}

        if not col_names:
            continue

        raw_map = dict(zip(col_names, col_raw)) if len(col_names) == len(col_raw) else {}
        num_cols = [c for c in col_names if col_types.get(c) == "numeric"]
        txt_cols = [c for c in col_names if col_types.get(c) == "text"]
        cat_cols = [c for c in txt_cols
                    if col_stats.get(c, {}).get("unique", 9999) < 50]

        hdr = f"\n[{fname}"
        if sheet and sheet != "CSV":
            hdr += f" | Foglio: {sheet}"
        hdr += f" | {sch['row_count']:,} righe]"
        parts.append(hdr)

        raw_col_list = ", ".join(raw_map.get(c, c) for c in col_names[:20])
        parts.append(f"Colonne: {raw_col_list}")

        if num_cols:
            parts.append("Statistiche numeriche:")
            for col in num_cols[:10]:
                s = col_stats.get(col, {})
                if s:
                    lbl = raw_map.get(col, col)
                    parts.append(
                        f"  {lbl}: sum={s.get('sum', 0):,.2f} | "
                        f"min={s.get('min', 0):,.2f} | max={s.get('max', 0):,.2f} | "
                        f"n={s.get('count', 0)}"
                    )

        if color_info:
            parts.append("Pattern colori (significato da inferire dal contesto del file):")
            for col, colors in list(color_info.items())[:5]:
                lbl = raw_map.get(col, col)
                for rgb, count in sorted(colors.items(), key=lambda x: -x[1])[:4]:
                    hex_c = f"#{rgb[2:] if len(rgb) == 8 else rgb}"
                    parts.append(f"  {lbl}: {count} celle con sfondo {hex_c}")

        if tname not in routed_tables:
            for cat_col in cat_cols[:2]:
                for num_col in num_cols[:2]:
                    rows = query_db(GENERIC_DB, f"""
                        SELECT "{cat_col}",
                               COUNT(*)         AS righe,
                               SUM("{num_col}") AS totale,
                               AVG("{num_col}") AS media
                        FROM "{tname}"
                        WHERE "{cat_col}" IS NOT NULL AND "{cat_col}" != ''
                        GROUP BY "{cat_col}"
                        ORDER BY totale DESC
                        LIMIT 25
                    """)
                    if rows and "error" not in rows[0] and len(rows) > 1:
                        cat_lbl = raw_map.get(cat_col, cat_col)
                        num_lbl = raw_map.get(num_col, num_col)
                        parts.append(f"\nBreakdown {cat_lbl} → {num_lbl}:")
                        for r in rows:
                            parts.append(
                                f"  {r[cat_col]}: "
                                f"totale={r['totale'] or 0:,.2f} | n={r['righe']} | "
                                f"media={r['media'] or 0:,.2f}"
                            )

        sample = query_db(GENERIC_DB, f'SELECT * FROM "{tname}" LIMIT 5')
        if sample and "error" not in sample[0]:
            parts.append("Righe di esempio:")
            for row in sample:
                line = " | ".join(
                    f"{raw_map.get(k, k)}={v}"
                    for k, v in list(row.items())[:8]
                    if v is not None and v != ""
                )
                parts.append(f"  {line}")

    return "\n".join(parts)


# ── PDF → Markdown (sync) ─────────────────────────────────────────────────────

def read_pdf_as_markdown(filepath) -> str:
    """PDF → Markdown con fallback markitdown → fitz → pdfplumber."""
    p = Path(filepath)
    header = f"=== PDF: {p.name} ===\n"

    try:
        from markitdown import MarkItDown
        result = MarkItDown().convert(str(filepath))
        text = result.text_content.strip()
        if len(text) > 100:
            return header + "[Formato: Markdown]\n\n" + text
    except ImportError:
        pass
    except Exception as e:
        print(f"  [markitdown warning] {p.name}: {e}", flush=True)

    fitz_text = ""
    try:
        doc = fitz.open(filepath)
        for i, page in enumerate(doc):
            t = page.get_text()
            if t.strip():
                fitz_text += f"[Pagina {i + 1}]\n{t}\n"
        doc.close()
    except Exception:
        pass

    plumber_text = ""
    try:
        import pdfplumber
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                for table in page.extract_tables() or []:
                    for row in table:
                        r = [str(c).strip() if c else "" for c in row]
                        if any(r):
                            plumber_text += " | ".join(r) + "\n"
    except Exception:
        pass

    parts = []
    if fitz_text:
        parts.append("[Formato: testo grezzo]\n" + fitz_text)
    if plumber_text:
        parts.append("[Tabelle estratte]\n" + plumber_text)

    return header + "\n".join(parts) if parts else f"[Impossibile estrarre {p.name}]"


# ── Memoria persistente ───────────────────────────────────────────────────────

def load_memory() -> dict:
    if MEMORY_FILE.exists():
        try:
            return json.loads(MEMORY_FILE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def save_memory(memory: dict):
    MEMORY_FILE.parent.mkdir(parents=True, exist_ok=True)
    MEMORY_FILE.write_text(
        json.dumps(memory, indent=2, ensure_ascii=False), encoding="utf-8"
    )


# ── Chunking + indexing (sync) ────────────────────────────────────────────────

def chunk_text(text: str) -> list:
    words = text.split()
    chunks, i = [], 0
    while i < len(words):
        chunks.append(" ".join(words[i:i + CHUNK_SIZE]))
        i += CHUNK_SIZE - CHUNK_OVERLAP
    return chunks or [""]


def index_file(filepath) -> int:
    """Indicizza un file su ChromaDB. Ritorna il numero di chunk creati."""
    p = Path(filepath)
    ext = p.suffix.lower()

    if ext not in EXTENSIONS["financial"]:
        return 0

    if ext in (".xlsx", ".xls"):
        try:
            sheets_data, text = read_excel_smart(filepath)
            if sheets_data:
                tables = build_file_db(filepath, sheets_data)
                rows = sum(r for _, _, r in tables)
                text += f"\n[SQLite: {rows:,} righe in {len(tables)} tabella/e]"
        except Exception as e:
            text = f"[Excel error: {e}]"

    elif ext == ".csv":
        try:
            sheets_data, text = read_csv_smart(filepath)
            if sheets_data:
                tables = build_file_db(filepath, sheets_data)
                rows = sum(r for _, _, r in tables)
                text += f"\n[SQLite: {rows:,} righe]"
        except Exception as e:
            text = f"[CSV error: {e}]"

    elif ext == ".pdf":
        text = read_pdf_as_markdown(filepath)

    elif ext == ".txt":
        text = open(filepath, encoding="utf-8", errors="ignore").read()

    else:
        return 0

    if not text or not text.strip():
        return 0

    try:
        existing = collection.get(where={"path": str(filepath)})
        if existing and existing["ids"]:
            collection.delete(ids=existing["ids"])
    except Exception:
        pass

    chunks = chunk_text(text)
    for i, chunk in enumerate(chunks):
        collection.upsert(
            documents=[chunk],
            ids=[f"{filepath}__c{i}"],
            metadatas=[{
                "filename": p.name,
                "path":     str(filepath),
                "chunk":    i,
                "agent":    "financial",
                "type":     "chunk",
            }]
        )
    return len(chunks)


def index_folder(folder):
    files = [f for f in Path(folder).rglob("*")
             if f.is_file() and f.suffix.lower() in EXTENSIONS["financial"]]
    print(f"[Financial] Trovati {len(files)} file...")
    total = 0
    for f in files:
        n = index_file(str(f))
        if n: print(f"  [OK] {f.name} → {n} chunk")
        else: print(f"  [--] {f.name} → saltato")
        total += n
    print(f"[Financial] Completato — {total} chunk totali")


# ── Search (async) ────────────────────────────────────────────────────────────

async def search(query: str) -> str:
    """Ricerca async con generazione lazy delle schede semantiche."""
    return await analyzer.search_with_cards(collection, query, "financial", n_results=4)


# ── System prompt ─────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """Sei un esperto analista finanziario aziendale con accesso diretto ai dati strutturati estratti dai file.

REGOLE ASSOLUTE:
- Rispondi SEMPRE e SOLO in italiano, qualunque sia la lingua della domanda o dei dati
- Usa ESCLUSIVAMENTE i dati forniti nel contesto — mai inventare, stimare o integrare con conoscenze esterne
- Riporta TUTTI i dati presenti nell'ordine in cui compaiono nel contesto
- Non scartare mai righe che sembrano anomale: se sono nel contesto sono dati reali e verificati
- Se un dato non è disponibile nel contesto, dichiaralo esplicitamente
- Cita sempre il file e il foglio sorgente quando disponibili

FORMATTAZIONE NUMERI:
- Separatore migliaia: punto  (es: 1.234.567)
- Separatore decimali: virgola (es: 1.234,56)
- Percentuali: 15,3%
- Valute: € 1.234,56

COLORI NELLE CELLE EXCEL:
- I colori delle celle non hanno un significato predefinito universale
- Interpreta il loro significato DAL CONTESTO SPECIFICO del documento in esame
  (es: se la colonna si chiama "Stato" e ha celle verdi/rosse, inferisci il significato)
- Descrivi i pattern rilevati e proponi un'interpretazione contestuale motivata
- Non assumere mai un significato fisso per nessun colore senza evidenza nel file

STRUTTURA MULTI-FOGLIO:
- Fogli diversi possono rappresentare periodi, categorie, o viste diverse dello stesso dataset
- Cerca relazioni tra fogli quando rilevante per la domanda
- Segnala se i dati richiesti sono distribuiti su più fogli

CAPACITÀ:
- Qualsiasi file finanziario: fatture, pagamenti, budget, bilanci, report di mercato, dashboard
- Aggregazioni precise: somme, medie, ranking, breakdown per categoria
- Confronti temporali, trend, analisi per segmento o canale
- Interpretazione di strutture dati complesse con fogli multipli e celle colorate"""


# ── User prompt builder (async per via di adaptive_sql_context) ───────────────

async def _build_user_prompt(question: str, context: str) -> str:
    loop = asyncio.get_running_loop()

    schemas = await loop.run_in_executor(None, get_all_file_schemas)
    generic_data = await adaptive_sql_context(question, schemas) if schemas else ""

    db_info = ""
    if schemas:
        db_info = f"[File DB: {len(schemas)} tabelle indicizzate]\n"
        for s in schemas[:6]:
            raw = s.get("_col_raw") or []
            db_info += (
                f"  {s['filename']} | {s['sheet']} "
                f"({s['row_count']:,} righe) — "
                f"{', '.join(str(c) for c in raw[:12])}\n"
            )

    memory = await loop.run_in_executor(None, load_memory)
    mem_txt = ""
    if memory:
        mem_txt = "\nMEMORIA:\n" + "\n".join(f"  {k}: {v}" for k, v in memory.items()) + "\n"

    has_db_data = bool(generic_data and generic_data.strip())
    raw_context = (
        "[Contesto semantico disponibile — usa i dati strutturati sopra per i valori numerici]"
        if has_db_data
        else (context[:3000] if len(context) > 3000 else context)
    )

    return f"""{mem_txt}
=== DATABASE ATTIVI ===
{db_info or "Nessun database indicizzato"}

=== DATI STRUTTURATI — FONTE PRIMARIA (usa SOLO questi per i valori numerici) ===
{generic_data or "—"}

=== CONTESTO SEMANTICO (solo per comprensione del documento, NON per i numeri) ===
{raw_context}

DOMANDA: {question}

RISPOSTA (obbligatoriamente in italiano — valori numerici solo dalla FONTE PRIMARIA):"""


# ── Answer (async, non-streaming) ─────────────────────────────────────────────

async def answer(question: str, context: str) -> str:
    user_prompt = await _build_user_prompt(question, context)
    return await chat_complete(
        model=ANSWER_MODEL,
        system=SYSTEM_PROMPT,
        user=user_prompt,
        temperature=0.2,
    )


# ── Answer streaming (async iterator) ─────────────────────────────────────────

async def answer_stream(question: str, context: str):
    user_prompt = await _build_user_prompt(question, context)
    async for chunk in chat_complete_stream(
        model=ANSWER_MODEL,
        system=SYSTEM_PROMPT,
        user=user_prompt,
        temperature=0.2,
    ):
        yield chunk


if __name__ == "__main__":
    folder = sys.argv[1] if len(sys.argv) > 1 else "."
    index_folder(folder)