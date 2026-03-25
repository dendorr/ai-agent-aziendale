import sys
import os
sys.path.append(os.path.expanduser("~/ai-agent"))

from config.config import WATCH_FOLDER, CHROMA_DB_PATH, SUPPORTED_EXTENSIONS
import chromadb
from pathlib import Path
import pdfplumber
import openpyxl
import ezdxf

# Inizializza ChromaDB
client = chromadb.PersistentClient(path=CHROMA_DB_PATH)
collection = client.get_or_create_collection("documenti_aziendali")

def leggi_pdf(filepath):
    testo = ""
    with pdfplumber.open(filepath) as pdf:
        for pagina in pdf.pages:
            t = pagina.extract_text()
            if t:
                testo += t + "\n"
    return testo

def leggi_excel(filepath):
    testo = ""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    for foglio in wb.sheetnames:
        ws = wb[foglio]
        testo += f"Foglio: {foglio}\n"
        for riga in ws.iter_rows(values_only=True):
            riga_pulita = [str(c) for c in riga if c is not None]
            if riga_pulita:
                testo += " | ".join(riga_pulita) + "\n"
    return testo

def leggi_pdf(filepath):
    import fitz
    testo = ""
    doc = fitz.open(filepath)
    for pagina in doc:
        testo += pagina.get_text() + "\n"
    doc.close()
    return testo

def leggi_file(filepath):
    ext = Path(filepath).suffix.lower()
    try:
        if ext == ".pdf":
            return leggi_pdf(filepath)
        elif ext in [".xlsx", ".xls"]:
            return leggi_excel(filepath)
        elif ext == ".dxf":
            return leggi_dxf(filepath)
        elif ext in [".txt", ".md", ".csv", ".svg"]:
            with open(filepath, "r", errors="ignore") as f:
                return f.read()
        else:
            return f"File {ext} rilevato ma estrazione testo non supportata."
    except Exception as e:
        return f"Errore nella lettura: {e}"

def chunk_testo(testo, chunk_size=500, overlap=50):
    parole = testo.split()
    chunks = []
    i = 0
    while i < len(parole):
        chunk = " ".join(parole[i:i+chunk_size])
        chunks.append(chunk)
        i += chunk_size - overlap
    return chunks

def indicizza_cartella(cartella):
    cartella = Path(cartella)
    files = [f for f in cartella.rglob("*") if f.suffix.lower() in SUPPORTED_EXTENSIONS]
    print(f"Trovati {len(files)} file da indicizzare...")
    for filepath in files:
        print(f"  Indicizzando: {filepath.name}")
        testo = leggi_file(str(filepath))
        if testo:
            chunks = chunk_testo(testo)
            for i, chunk in enumerate(chunks):
                collection.upsert(
                    documents=[chunk],
                    ids=[f"{str(filepath)}__chunk{i}"],
                    metadatas=[{"filename": filepath.name, "path": str(filepath), "chunk": i}]
                )
            print(f"    → {len(chunks)} chunks creati")
    print("Indicizzazione completata!")

if __name__ == "__main__":
    print(f"Cartella monitorata: {WATCH_FOLDER}")
    indicizza_cartella(WATCH_FOLDER)
