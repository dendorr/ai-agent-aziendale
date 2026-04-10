"""
NIELSEN DB BUILDER v2
Auto-detects header structure (3 or 4 rows) for each Share sheet.
Handles all segment types: NCC, RG, NDG, LAMM, Pods, Beans, Instant, etc.
"""
import sys, os, sqlite3
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from pathlib import Path

DB_PATH = Path("~/ai-agent/memory/nielsen.db").expanduser()

# Nielsen standard aggregate brand codes — not real brands, exclude from rankings.
# AO = All Others, AO_ variants = All Others in sub-group.
AGGREGATE_BRANDS = {"AO", "ALL OTHERS", "OTHERS", "OTHER"}

SEGMENT_MAP = {
    "Share NCC":                  "NCC",
    "Share NCC Decaf":            "NCC_Decaf",
    "Share R&G":                  "RG",
    "Share R&G Decaff":           "RG_Decaf",
    "Share NDG":                  "NDG",
    "Share LAMM":                 "LAMM",
    "SharePods":                  "Pods",
    "SharePods Decaf":            "Pods_Decaf",
    "Share Beans":                "Beans",
    "Share Instant tot":          "Instant_Tot",
    "Share Instant pure soluble": "Instant_Pure",
    "Share Instant pure decaf":   "Instant_Decaf",
    "Share Instant MIxes":        "Instant_Mixes",
    "Share Instant Specialties":  "Instant_Spec",
    "Share Tot Coffee":           "Tot_Coffee",
    "Canali":                     "Tot_Coffee_Canali",
    "Segmenti":                   "Segmenti",
}

METRIC_LABELS = {
    "VOLUME":               "VOLUME_KGS",
    "VALUE":                "VALUE_EUR",
    "VOLUME SOM":           "VOLUME_SOM",
    "VALUE SOM":            "VALUE_SOM",
    "€/KG":                 "PRICE_PER_KG",
    "PINDEX €/kg":          "PRICE_INDEX",
    "PINDEX €/CUP":         "PRICE_INDEX_CUP",
    "€/CUP":                "PRICE_PER_CUP",
    "WEIGHTED DISTRIBUTION":"WEIGHTED_DIST",
    "CUP SOM":              "CUP_SOM",
    "VALUE SOM":            "VALUE_SOM",
}

PERIOD_LABELS = {"LM12 -1","L12M","L6M -1","L6M","L3M -1","L3M",
                 "LM-1","LM","YTD -1","YTD","Vs PY %"}

# Market prefixes ordered longest-first to avoid "Food" matching before "Food A1"
MARKET_PREFIXES = [
    "S+H Lombardia","S+H Lazio","S+H Campania","S+H Sicilia","S+H Veneto",
    "S+H Piemonte","S+H Emilia Romagna","S+H Toscana","S+H Puglia",
    "S+H Calabria","S+H Sardegna","S+H Abruzzo","S+H Basilicata",
    "S+H Molise","S+H Liguria","S+H FVG","S+H TAA","S+H Umbria",
    "S+H Marche","S+H Valle D'Aosta",
    "Hyp+Sup+SS Area 1","Hyp+Sup+SS Area 2","Hyp+Sup+SS Area 3","Hyp+Sup+SS Area 4",
    "Hyper + Super Area 1","Hyper + Super Area 2","Hyper + Super Area 3","Hyper + Super Area 4",
    "Hyp+Sup Abruzzo","Hyp+Sup Basilicata","Hyp+Sup Calabria","Hyp+Sup Campania",
    "Hyp+Sup Emilia Romagna","Hyp+Sup Friuli Venezia Giulia","Hyp+Sup Lazio",
    "Hyp+Sup Liguria","Hyp+Sup Lombardia","Hyp+Sup Marche","Hyp+Sup Molise",
    "Hyp+Sup Piemonte","Hyp+Sup Puglia","Hyp+Sup Sardegna","Hyp+Sup Sicilia",
    "Hyp+Sup Toscana","Hyp+Sup Trentino Alto Adige","Hyp+Sup Umbria",
    "Hyp+Sup Valle DAosta","Hyp+Sup Veneto",
    "Hyper 2500-4999","Hyper 5000-7999","Hyper >8000",
    "Hyp+Sup+SS","Hyper + Super","Hyper",
    "S+H A1","S+H A2","S+H A3","S+H A4","S+H",
    "Super 1000-2499","Super 400-999","Super",
    "Food A1","Food A2","Food A3","Food A4",
    "Modern Distribution","Omnichannel","E-Commerce",
    "Discount","Discounters","Superettes","Food",
    "Area 1","Area 2","Area 3","Area 4",
    "COOP",
]

def is_formula(v): return isinstance(v, str) and str(v).strip().startswith("=")
def clean(v):
    if v is None or is_formula(v): return None
    s = str(v).strip()
    return s if s else None

def extract_market(composite_key, brand):
    """Extract market from composite key like 'Food A1NESPRESSO COMPATIBLES'."""
    if not composite_key: return "Unknown"
    key = str(composite_key).strip()
    for prefix in MARKET_PREFIXES:
        if key.startswith(prefix):
            return prefix
    # Fallback: try removing the brand from the key
    if brand and brand in key:
        mkt = key.replace(brand, "").strip()
        if mkt: return mkt
    return key[:30]

def detect_sheet_structure(all_rows):
    """
    Auto-detect header rows.
    Returns (metric_row_idx, period_row_idx, data_start_idx, metric_cols)
    """
    metric_row_idx  = None
    period_row_idx  = None
    metric_cols     = {}

    for row_i, row in enumerate(all_rows[:6]):
        if not row: continue
        # Check if this row has metric group labels
        found_metrics = {}
        for col_i, val in enumerate(row):
            c = clean(val)
            if c and c in METRIC_LABELS:
                found_metrics[col_i] = METRIC_LABELS[c]
        if found_metrics:
            metric_row_idx = row_i
            metric_cols    = found_metrics
            continue
        # Check if this row has period labels
        if metric_row_idx is not None:
            found_periods = sum(1 for v in row if clean(v) in PERIOD_LABELS)
            if found_periods >= 3:
                period_row_idx = row_i
                break

    if metric_row_idx is None or period_row_idx is None:
        # Fallback: look for rows with period labels directly
        for row_i, row in enumerate(all_rows[:6]):
            found_periods = sum(1 for v in row if clean(v) in PERIOD_LABELS)
            if found_periods >= 3:
                period_row_idx = row_i
                # metric row is the previous one
                if row_i > 0 and not metric_cols:
                    for col_i, val in enumerate(all_rows[row_i - 1]):
                        c = clean(val)
                        if c and c in METRIC_LABELS:
                            metric_cols[col_i] = METRIC_LABELS[c]
                    if metric_cols:
                        metric_row_idx = row_i - 1
                break

    data_start = (period_row_idx + 1) if period_row_idx is not None else 4
    return metric_row_idx, period_row_idx, data_start, metric_cols

def parse_period_headers(row, metric_cols):
    """
    Build a map of col_index -> (metric, period) from the period header row.
    Metric group for a column = closest metric_col to the left.
    """
    col_map = {}
    sorted_metric_cols = sorted(metric_cols.keys())

    for col_i, val in enumerate(row):
        period = clean(val)
        if not period or period not in PERIOD_LABELS or period == "Vs PY %":
            continue
        # Find which metric group this period belongs to
        metric = None
        for mc in reversed(sorted_metric_cols):
            if mc <= col_i:
                metric = metric_cols[mc]
                break
        if metric:
            col_map[col_i] = (metric, period)

    return col_map

def hierarchy_level(hierarchy_path):
    """
    Derive the depth level from the hierarchy path (col 3).
      0 = category aggregate  (e.g. "NESPRESSO COMPATIBLES")
      1 = direct brand        (e.g. "NESPRESSO COMPATIBLES | LAVAZZA")
      2 = sub-brand           (e.g. "NESPRESSO COMPATIBLES | JDE | L'OR")
    """
    if not hierarchy_path:
        return 0
    return hierarchy_path.count("|")

def process_sheet(ws, segment):
    """
    Process one Share sheet.
    Returns list of (segment, brand, market, metric, period, value, hierarchy, level).
    """
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 3: return []

    _, period_row_idx, data_start, metric_cols = detect_sheet_structure(all_rows)

    if period_row_idx is None or not metric_cols:
        return []

    period_row = all_rows[period_row_idx]
    col_map    = parse_period_headers(period_row, metric_cols)

    if not col_map: return []

    records = []
    for row in all_rows[data_start:]:
        if not row: continue

        composite = clean(row[0]) if len(row) > 0 else None

        # Col 3 = full hierarchy path (e.g. "NESPRESSO COMPATIBLES | JDE | L'OR")
        hierarchy = clean(row[3]) if len(row) > 3 else None

        # Col 4 = last element of hierarchy = the brand at this level
        brand = None
        for try_col in [4, 2, 1]:
            if len(row) > try_col:
                v = clean(row[try_col])
                if v and "|" not in v and not v.isdigit():
                    is_market = any(v.startswith(p) for p in
                                    ["Area", "Food", "Hyper", "Super", "Disc", "Omni"])
                    if not is_market:
                        brand = v
                        break

        if not brand or not composite: continue

        market       = extract_market(composite, brand)
        level        = hierarchy_level(hierarchy)
        is_aggregate = 1 if brand.upper() in AGGREGATE_BRANDS or level == 0 else 0

        for col_i, (metric, period) in col_map.items():
            if col_i >= len(row): continue
            val = row[col_i]
            if val is None or is_formula(val): continue
            try:
                num = float(val)
                records.append((segment, brand, market, metric, period, num,
                                hierarchy or brand, level, is_aggregate))
            except (ValueError, TypeError):
                continue

    return records

def build_db(excel_path):
    import openpyxl
    print(f"Loading {Path(excel_path).name}...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    if DB_PATH.exists(): DB_PATH.unlink()

    conn   = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE nielsen_data (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            segment      TEXT, brand      TEXT, market TEXT,
            metric       TEXT, period     TEXT, value  REAL,
            hierarchy    TEXT,   -- full path, e.g. "NCC | LAVAZZA"
            level        INTEGER, -- 0=category aggregate, 1=brand, 2=sub-brand
            is_aggregate INTEGER  -- 1 if AO/ALL OTHERS/category total, 0 if real brand
        )
    """)
    cursor.execute("CREATE TABLE file_info (key TEXT PRIMARY KEY, value TEXT)")
    cursor.execute("INSERT INTO file_info VALUES ('source', ?)", (str(excel_path),))

    total = 0; sheets_ok = 0
    for sheet_name, segment in SEGMENT_MAP.items():
        if sheet_name not in wb.sheetnames: continue
        ws      = wb[sheet_name]
        records = process_sheet(ws, segment)
        if records:
            cursor.executemany(
                "INSERT INTO nielsen_data (segment,brand,market,metric,period,value,hierarchy,level,is_aggregate) VALUES (?,?,?,?,?,?,?,?,?)",
                records
            )
            total     += len(records)
            sheets_ok += 1
            print(f"  [OK] {sheet_name:35s} -> {len(records):6,} records")
        else:
            print(f"  [--] {sheet_name:35s} -> 0 records (structure not recognized)")

    # Indexes
    for col in ("segment","brand","market","metric","period","level","is_aggregate"):
        cursor.execute(f"CREATE INDEX idx_{col} ON nielsen_data({col})")
    conn.commit()
    conn.close()
    print(f"\nDB built: {total:,} records from {sheets_ok} sheets → {DB_PATH}")
    return DB_PATH

def query_db(sql, params=()):
    if not DB_PATH.exists(): return []
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        return [dict(r) for r in conn.execute(sql, params).fetchall()]
    except Exception as e:
        return [{"error": str(e)}]
    finally:
        conn.close()

def get_db_summary():
    if not DB_PATH.exists(): return "DB non trovato."
    total   = query_db("SELECT COUNT(*) n FROM nielsen_data")[0]["n"]
    brands  = query_db("SELECT COUNT(DISTINCT brand)  n FROM nielsen_data")[0]["n"]
    markets = query_db("SELECT COUNT(DISTINCT market) n FROM nielsen_data")[0]["n"]
    segs    = query_db("SELECT segment, COUNT(*) n FROM nielsen_data GROUP BY segment ORDER BY n DESC")
    periods = query_db("SELECT DISTINCT period FROM nielsen_data ORDER BY period")
    txt  = f"Database Nielsen: {total:,} record\n"
    txt += f"Brand unici: {brands} | Mercati: {markets}\n"
    txt += "Segmenti:\n" + "\n".join(f"  {r['segment']}: {r['n']:,}" for r in segs) + "\n"
    txt += f"Periodi: {', '.join(r['period'] for r in periods)}\n"
    return txt

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python nielsen_db_builder.py file.xlsx")
        sys.exit(1)
    build_db(sys.argv[1])
    print("\n" + get_db_summary())